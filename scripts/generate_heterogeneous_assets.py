#!/usr/bin/env python3
"""Generate KWOK node manifests, optional simulator overrides, and a compositional hardware catalog.

Input columns expected (case-insensitive):
- node_name
- vendor (nvidia|amd|none)
- product
- cpu
- cpu_sockets
- memory_gib
- gpu_count
- gpu_min_cap_watts
- gpu_max_cap_watts

Usage:
  python scripts/generate_heterogeneous_assets.py \
    --input inventory.xlsx \
    --sheet Nodes \
    --out-nodes "examples/07 - simulator-gpu-powercaps/manifests/00-kwok-nodes.yaml" \
    --out-classes "examples/07 - simulator-gpu-powercaps/manifests/10-node-classes.yaml" \
    --out-catalog simulator/catalog/hardware.generated.yaml
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import OrderedDict
from pathlib import Path


def _strip_yaml_scalar(value: str) -> str:
    value = value.strip()
    if (value.startswith('"') and value.endswith('"')) or (value.startswith("'") and value.endswith("'")):
        return value[1:-1]
    return value


def load_simple_yaml_rows(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    current: dict[str, str] | None = None
    in_nodes = False
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.split("#", 1)[0].rstrip()
        if not line.strip():
            continue
        stripped = line.strip()
        if stripped == "nodes:":
            in_nodes = True
            continue
        if not in_nodes:
            continue
        if stripped.startswith("- "):
            if current:
                rows.append(current)
            current = {}
            item = stripped[2:].strip()
            if item and ":" in item:
                key, value = item.split(":", 1)
                current[key.strip()] = _strip_yaml_scalar(value)
            continue
        if current is None:
            continue
        if ":" not in stripped:
            continue
        key, value = stripped.split(":", 1)
        current[key.strip()] = _strip_yaml_scalar(value)
    if current:
        rows.append(current)
    return rows


def load_rows(path: Path, sheet: str | None) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix in {".csv"}:
        with path.open(newline="", encoding="utf-8") as f:
            return list(csv.DictReader(f))
    if suffix in {".yaml", ".yml"}:
        try:
            import yaml  # type: ignore
        except Exception as exc:  # pragma: no cover
            rows = load_simple_yaml_rows(path)
            if not rows:
                raise RuntimeError("pyyaml is required for general .yaml input, or provide a simple nodes: list format") from exc
            return rows
        else:
            with path.open(encoding="utf-8") as f:
                data = yaml.safe_load(f)
            if isinstance(data, dict) and isinstance(data.get("nodes"), list):
                rows = data["nodes"]
            elif isinstance(data, list):
                rows = data
            else:
                raise RuntimeError("yaml input must be a list of node rows or a mapping with a 'nodes' list")
            out: list[dict[str, str]] = []
            for row in rows:
                if not isinstance(row, dict):
                    continue
                out.append({str(k): "" if v is None else str(v).strip() for k, v in row.items()})
            return out
    if suffix in {".xlsx", ".xlsm"}:
        try:
            import openpyxl  # type: ignore
        except Exception as exc:  # pragma: no cover
            raise RuntimeError("openpyxl is required for .xlsx input") from exc
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        header = [str(c.value or "").strip() for c in ws[1]]
        rows: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            rows.append({header[i]: "" if row[i] is None else str(row[i]).strip() for i in range(len(header))})
        return rows
    raise RuntimeError(f"unsupported input format: {path.suffix}")


def norm(row: dict[str, str], key: str, default: str = "") -> str:
    for k, v in row.items():
        if k.strip().lower() == key.lower():
            return (v or "").strip()
    return default


def first_nonempty(row: dict[str, str], *keys: str, default: str = "") -> str:
    for key in keys:
        val = norm(row, key)
        if val:
            return val
    return default


def parse_intish(value: str, default: int = 0) -> int:
    if value is None:
        return default
    txt = str(value).strip()
    if not txt:
        return default
    txt = txt.replace(",", "")
    try:
        return int(float(txt))
    except ValueError:
        return default


def parse_memory_gib(value: str, default: int = 64) -> int:
    txt = str(value or "").strip().upper().replace(" ", "")
    if not txt:
        return default
    m = re.match(r"^([0-9]+(?:\.[0-9]+)?)([TGM]I?B?)?$", txt)
    if not m:
        return default
    qty = float(m.group(1))
    unit = (m.group(2) or "G").upper()
    if unit.startswith("T"):
        qty *= 1024
    elif unit.startswith("M"):
        qty /= 1024
    return max(1, int(round(qty)))


def slugify(value: str) -> str:
    out = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return out or "node"


def label_safe(value: str) -> str:
    txt = re.sub(r"[^A-Za-z0-9_.-]+", "-", str(value or "").strip()).strip("-.")
    return txt or "unknown"


def yaml_list(items: list[str]) -> str:
    quoted = [f'"{item}"' for item in items if item]
    return "[" + ", ".join(quoted) + "]"


def expand_inventory_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    expanded: list[dict[str, str]] = []
    for row in rows:
        replicas = parse_intish(first_nonempty(row, "replicas", "nodes", default="1"), 1)
        node_name = norm(row, "node_name")
        node_name_prefix = first_nonempty(row, "node_name_prefix", "node/type", "description")
        if not node_name and node_name_prefix:
            node_name_prefix = slugify(node_name_prefix)
        elif node_name and not node_name_prefix:
            node_name_prefix = re.sub(r"-\d+$", "", node_name)
        if not node_name and not node_name_prefix:
            node_name_prefix = "kwok-node"
        for idx in range(replicas):
            item = dict(row)
            if not node_name:
                item["node_name"] = f"{node_name_prefix}-{idx}"
            elif replicas == 1:
                item["node_name"] = node_name
            else:
                item["node_name"] = f"{node_name_prefix}-{idx}"
            expanded.append(item)
    return expanded


def normalize_inventory_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    normalized: list[dict[str, str]] = []
    for raw in rows:
        product = first_nonempty(raw, "product", "nvidia.com/gpu.product")
        cpu_model = first_nonempty(raw, "cpu_model", "cpu")
        cpu_sockets = first_nonempty(raw, "cpu_sockets", "sockets", default="2")
        cpu_cores = first_nonempty(raw, "cpu_cores", "cores", default="")
        memory_gib = first_nonempty(raw, "memory_gib", "ram")
        gpu_count = first_nonempty(raw, "gpu_count", "gpu", default="0")
        vendor = norm(raw, "vendor", "").lower()
        if not vendor:
            if product:
                vendor = "amd" if "amd" in product.lower() else "nvidia"
            else:
                vendor = "none"
        normalized.append(
            {
                **raw,
                "vendor": vendor,
                "product": product,
                "cpu": cpu_model,
                "cpu_sockets": str(parse_intish(cpu_sockets, 2)),
                "cpu_cores": str(parse_intish(cpu_cores, 16)),
                "memory_gib": str(parse_memory_gib(memory_gib, 64)),
                "gpu_count": str(parse_intish(gpu_count, 0)),
                "gpu_min_cap_watts": first_nonempty(raw, "gpu_min_cap_watts", default=""),
                "gpu_max_cap_watts": first_nonempty(raw, "gpu_max_cap_watts", default=""),
            }
        )
    return expand_inventory_rows(normalized)


def make_kwok_node(row: dict[str, str]) -> str:
    name = norm(row, "node_name")
    vendor = norm(row, "vendor", "none").lower()
    product = norm(row, "product")
    cpu_model = norm(row, "cpu")
    cpu_model_label = label_safe(cpu_model)
    gpu_model_label = label_safe(product)
    cpu_cores = norm(row, "cpu_cores", "16")
    mem = norm(row, "memory_gib", "64")
    gpu_count = norm(row, "gpu_count", "0")
    has_gpu = gpu_count not in {"", "0"}

    labels = [
        '    type: kwok',
        '    joulie.io/managed: "true"',
        f'    joulie.io/node-name: "{name}"',
        f'    joulie.io/hw.cpu-model: "{cpu_model_label}"' if cpu_model else None,
        f'    joulie.io/hw.cpu-cores: "{cpu_cores}"',
        f'    joulie.io/hw.cpu-sockets: "{norm(row, "cpu_sockets", "2")}"',
        f'    joulie.io/hw.gpu-model: "{gpu_model_label}"' if product else None,
        f'    joulie.io/hw.gpu-count: "{gpu_count}"',
        f'    joulie.io/hw.kind: "{"gpu" if has_gpu else "cpu-only"}"',
        f'    joulie.io/gpu.product: "{gpu_model_label}"' if product else None,
    ]
    alloc_gpu_lines: list[str] = []
    if vendor == "nvidia" and gpu_count not in {"", "0"}:
        labels.append('    feature.node.kubernetes.io/pci-10de.present: "true"')
        alloc_gpu_lines = [f'    nvidia.com/gpu: "{gpu_count}"']
    elif vendor == "amd" and gpu_count not in {"", "0"}:
        labels.append('    feature.node.kubernetes.io/pci-1002.present: "true"')
        alloc_gpu_lines = [f'    amd.com/gpu: "{gpu_count}"']

    labels = [x for x in labels if x]
    block = [
        "apiVersion: v1",
        "kind: Node",
        "metadata:",
        f"  name: {name}",
        "  annotations:",
        "    kwok.x-k8s.io/node: fake",
        f'    joulie.io/raw.cpu-model: "{cpu_model}"' if cpu_model else None,
        f'    joulie.io/raw.gpu-product: "{product}"' if product else None,
        "  labels:",
        *labels,
        "spec:",
        "  taints:",
        "    - key: kwok.x-k8s.io/node",
        "      value: fake",
        "      effect: NoSchedule",
        "status:",
        "  allocatable:",
        f'    cpu: "{cpu_cores}"',
        *alloc_gpu_lines,
        f'    memory: "{mem}Gi"',
        '    pods: "110"',
        "  capacity:",
        f'    cpu: "{cpu_cores}"',
        *alloc_gpu_lines,
        f'    memory: "{mem}Gi"',
        '    pods: "110"',
    ]
    block = [line for line in block if line is not None]
    return "\n".join(block)


# Per-GPU-product hardware physics parameters.
#
# idleWattsPerGpu:
#   Board-level idle power divided by GPU count. Sources:
#   - H100 NVL: ~450-480W board idle / 8 GPUs ≈ 60W/GPU. NVLink fabric adds ~20-30W idle
#     overhead vs discrete cards (NVIDIA H100 NVL PCIe spec, board TBP 400W max, ~480W strap).
#   - H100 SXM5: SXM5 socket with NVLink bridge; field measurements report ~800W board idle
#     / 8 GPUs ≈ 100W/GPU (MLCommons MLPerf Training v4.1 power submissions).
#   - L40S: PCIe card; NVIDIA specifies 30W idle board power, ~35W with GDDR6 retention.
#   - MI300X: Unified memory APU-like design; AMD datasheet shows 700-750W TBP; memory
#     coherency circuits and HBM3 self-refresh drive ~120W/GPU idle floor.
#   - W7900: Workstation PCIe card; RDNA3 power floor ~25W/GPU (AMD Radeon PRO W7900 spec).
#
# computeGamma / memoryEpsilon / memoryGamma  (CappedBoardGPUModel parameters):
#   computeScale = (capW / natW)^computeGamma
#   memScale     = 1 - memoryEpsilon * (1 - capW/natW)^memoryGamma
#
#   Calibrated from:
#   - MLPerf Power submissions (v3.1/v4.0): H100 shows steep compute regression under capping;
#     MI300X compute holds up better due to unified-memory design.
#   - "Characterizing Power Management Opportunities on DGX H100" (NeurIPS'23 Systems Workshop):
#     H100 compute throughput drops ~30% at 80% power cap → gamma ≈ 1.4-1.5.
#   - AMD MI300X architecture whitepaper: unified memory preserves compute at lower clock floors;
#     CDNA3 sustains ~90% compute at 85% power cap → gamma ≈ 0.85.
#   - GDDR6 bandwidth vs power: GDDR6 (L40S, W7900) is more sensitive than HBM3 to clock
#     throttling → higher memoryEpsilon (0.25-0.30) vs HBM3 (0.10-0.15).
_GPU_PHYSICS: dict[str, dict] = {
    # NVIDIA H100 NVL (HBM3, NVLink, 400W max TBP per GPU)
    "NVIDIA H100 NVL": {
        "idleWattsPerGpu": 60,
        "computeGamma": 1.50,
        "memoryEpsilon": 0.15,
        "memoryGamma": 0.90,
    },
    # NVIDIA H100 80GB HBM3 (SXM5 socket, NVLink bridge, 700W max TBP per GPU)
    "NVIDIA H100 80GB HBM3": {
        "idleWattsPerGpu": 100,
        "computeGamma": 1.40,
        "memoryEpsilon": 0.15,
        "memoryGamma": 0.90,
    },
    # NVIDIA L40S (GDDR6, inference/rendering, 350W max TBP)
    "NVIDIA L40S": {
        "idleWattsPerGpu": 35,
        "computeGamma": 1.20,
        "memoryEpsilon": 0.25,
        "memoryGamma": 1.10,
    },
    # AMD Instinct MI300X (HBM3 unified memory, 750W max TBP)
    "AMD Instinct MI300X": {
        "idleWattsPerGpu": 120,
        "computeGamma": 0.85,
        "memoryEpsilon": 0.10,
        "memoryGamma": 0.85,
    },
    # AMD Radeon PRO W7900 (GDDR6, workstation, 295W max TBP)
    "AMD Radeon PRO W7900": {
        "idleWattsPerGpu": 25,
        "computeGamma": 1.10,
        "memoryEpsilon": 0.30,
        "memoryGamma": 1.20,
    },
}

_GPU_PHYSICS_DEFAULTS: dict = {
    "idleWattsPerGpu": 30,
    "computeGamma": 1.0,
    "memoryEpsilon": 0.20,
    "memoryGamma": 1.10,
}


def _gpu_physics(product: str) -> dict:
    """Return hardware physics params for a GPU product, falling back to defaults."""
    return _GPU_PHYSICS.get(product, _GPU_PHYSICS_DEFAULTS)


def make_class_rows(rows: list[dict[str, str]]) -> str:
    classes: OrderedDict[str, dict[str, str]] = OrderedDict()
    for row in rows:
        vendor = norm(row, "vendor", "none").lower()
        product = norm(row, "product")
        if vendor not in {"nvidia", "amd"}:
            continue
        key = f"{vendor}:{product}"
        if key in classes:
            continue
        classes[key] = {
            "vendor": vendor,
            "product": product,
            "count": norm(row, "gpu_count", "1"),
            "min": norm(row, "gpu_min_cap_watts", "200"),
            "max": norm(row, "gpu_max_cap_watts", "350"),
        }

    out = ["classes:"]
    for idx, (_, item) in enumerate(classes.items(), start=1):
        product_label = label_safe(item["product"])
        phys = _gpu_physics(item["product"])
        out.extend(
            [
                f"  - name: gpu-class-{idx}",
                "    matchLabels:",
                f'      joulie.io/gpu.product: "{product_label}"',
                "    model:",
                "      gpu:",
                f'        vendor: {item["vendor"]}',
                f'        product: "{item["product"]}"',
                f'        count: {item["count"]}',
                f'        idleWattsPerGpu: {phys["idleWattsPerGpu"]}',
                f'        maxWattsPerGpu: {item["max"]}',
                f'        minCapWattsPerGpu: {item["min"]}',
                "        powerModel:",
                "          alphaUtil: 1.0",
                "          betaCap: 1.0",
                f'        computeGamma: {phys["computeGamma"]}',
                f'        memoryEpsilon: {phys["memoryEpsilon"]}',
                f'        memoryGamma: {phys["memoryGamma"]}',
            ]
        )
    return "\n".join(out) + "\n"


def make_catalog_rows(rows: list[dict[str, str]]) -> str:
    cpu_models: OrderedDict[str, dict[str, str]] = OrderedDict()
    gpu_models: OrderedDict[str, dict[str, str]] = OrderedDict()
    out = ["catalogVersion: v1", "cpuModels:"]
    for row in rows:
        cpu = norm(row, "cpu")
        if not cpu or cpu in cpu_models:
            continue
        vendor = "amd" if "amd" in cpu.lower() else "intel" if "intel" in cpu.lower() or "xeon" in cpu.lower() else ""
        cpu_models[cpu] = {
            "vendor": vendor,
            "cores": norm(row, "cpu", "").split(" ")[-2] if "Core" in cpu else "",
        }
    for cpu, item in cpu_models.items():
        key = cpu.replace("-", "_").replace(" ", "_").replace("(R)", "").replace("__", "_")
        out.extend(
            [
                f"  {key}:",
                f"    aliases: {yaml_list([cpu, label_safe(cpu)])}",
                "    provenance: generated-from-inventory",
                "    official:",
                f'      vendor: {item["vendor"] or "unknown"}',
                "      baseGHz: 2.0",
                "      boostGHz: 3.0",
                "      tdpW: 300",
            ]
        )
    out.append("")
    out.append("gpuModels:")
    for row in rows:
        vendor = norm(row, "vendor", "none").lower()
        product = norm(row, "product")
        if vendor not in {"nvidia", "amd"} or not product:
            continue
        if product in gpu_models:
            continue
        gpu_models[product] = {
            "vendor": vendor,
            "max": norm(row, "gpu_max_cap_watts", "350"),
        }
    for product, item in gpu_models.items():
        out.extend(
            [
                f"  {product.replace('-', '_').replace(' ', '_')}:",
                f"    aliases: {yaml_list([product, label_safe(product)])}",
                "    provenance: generated-from-inventory",
                "    official:",
                f'      vendor: {item["vendor"]}',
                f'      maxBoardPowerW: {item["max"]}',
            ]
        )
    return "\n".join(out) + "\n"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, type=Path)
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--out-nodes", required=True, type=Path)
    ap.add_argument("--out-classes", required=True, type=Path)
    ap.add_argument("--out-catalog", type=Path, default=None)
    args = ap.parse_args()

    rows = normalize_inventory_rows(load_rows(args.input, args.sheet))
    if not rows:
        raise RuntimeError("no inventory rows found")

    nodes_yaml = "\n---\n".join(make_kwok_node(r) for r in rows) + "\n"
    classes_yaml = make_class_rows(rows)
    catalog_yaml = make_catalog_rows(rows)

    args.out_nodes.parent.mkdir(parents=True, exist_ok=True)
    args.out_classes.parent.mkdir(parents=True, exist_ok=True)
    args.out_nodes.write_text(nodes_yaml, encoding="utf-8")
    args.out_classes.write_text(classes_yaml, encoding="utf-8")
    if args.out_catalog:
        args.out_catalog.parent.mkdir(parents=True, exist_ok=True)
        args.out_catalog.write_text(catalog_yaml, encoding="utf-8")

    print(f"wrote {args.out_nodes}")
    print(f"wrote {args.out_classes}")
    if args.out_catalog:
        print(f"wrote {args.out_catalog}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # pragma: no cover
        print(f"error: {exc}", file=sys.stderr)
        raise SystemExit(1)
